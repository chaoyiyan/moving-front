from PIL import Image
import os
import openpyxl

def get_all_horizontal_gray_values(image_path):
    image = Image.open(image_path).convert('L')  # Convert the image to grayscale
    width, height = image.size
    gray_values = []

    topmost_row = int(height * 0.3)  # Calculate the topmost row position (30% of the height)

    for x in range(width):
        pixel_value = image.getpixel((x, topmost_row))  # Get the pixel value at the topmost row for each column
        gray_values.append(pixel_value)

    return gray_values

def process_images_in_folder(folder_path, output_excel_path):
    # Get a list of image file paths in the folder
    image_files = [file for file in os.listdir(folder_path) if file.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.gif'))]

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Horizontal Gray Values"

    col_counter = 1

    for image_file in image_files:
        image_path = os.path.join(folder_path, image_file)
        gray_values = get_all_horizontal_gray_values(image_path)

        worksheet.cell(row=1, column=col_counter, value=f"{image_file} (Position)")

        for col_idx, gray_value in enumerate(gray_values, start=2):
            worksheet.cell(row=col_idx, column=col_counter, value=col_idx - 1)  # Subtract 1 to get the relative position
            worksheet.cell(row=col_idx, column=col_counter + 1, value=gray_value)

        col_counter += 2  # Increment by 2 to leave space for the next image

    workbook.save(output_excel_path)
    print(f"Processed {len(image_files)} images and saved results to '{output_excel_path}'.")

if __name__ == "__main__":
    input_folder_path = r"D:\Codes\grey_value"
    output_excel_path = r"D:\Codes\grey_value\excel.xlsx"
    process_images_in_folder(input_folder_path, output_excel_path)
